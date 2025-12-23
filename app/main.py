from fastapi import FastAPI, File, UploadFile, HTTPException, Request
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
import pandas as pd
import openpyxl
from io import BytesIO
import os
from datetime import datetime
from .dns_analyzer import DNSAnalyzer
from .models import AnalysisResponse, AnalysisRequest
import uuid
import logging
import traceback
import re
from pathlib import Path

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Security: Rate limiting (simple in-memory implementation)
from collections import defaultdict
from time import time

class RateLimitMiddleware(BaseHTTPMiddleware):
    def __init__(self, app, requests_per_minute: int = 30):
        super().__init__(app)
        self.requests_per_minute = requests_per_minute
        self.clients = defaultdict(list)
    
    async def dispatch(self, request: Request, call_next):
        client_ip = request.client.host
        now = time()
        
        # Clean old requests (older than 1 minute)
        self.clients[client_ip] = [req_time for req_time in self.clients[client_ip] if now - req_time < 60]
        
        if len(self.clients[client_ip]) >= self.requests_per_minute:
            return HTMLResponse(
                content="<h1>429 Too Many Requests</h1><p>Rate limit exceeded. Please try again later.</p>",
                status_code=429
            )
        
        self.clients[client_ip].append(now)
        response = await call_next(request)
        return response

class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        # Add security headers
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['X-XSS-Protection'] = '1; mode=block'
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
        response.headers['Content-Security-Policy'] = "default-src 'self' https://cdnjs.cloudflare.com; style-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com; script-src 'self' 'unsafe-inline'; font-src 'self' https://cdnjs.cloudflare.com"
        return response

# Get the project root directory
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(PROJECT_ROOT, "static")
TEMPLATES_DIR = os.path.join(PROJECT_ROOT, "templates")

# Create FastAPI app
app = FastAPI(
    title="Domain Analyzer - Utopia Tech",
    description="Analyze DNS records for domains from uploaded Excel files. Powered by Utopia Tech (www.utopiats.com)",
    version="1.0.0"
)

# Configure multipart form size limits
app.state.max_file_size = 10 * 1024 * 1024  # 10MB

# Add security middleware
app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(RateLimitMiddleware, requests_per_minute=30)

# Add CORS middleware - SECURITY: Restrict origins in production
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # TODO: Restrict to specific origins in production
    allow_credentials=True,
    allow_methods=["GET", "POST"],  # Only allow needed methods
    allow_headers=["*"],
)

# Mount static files and templates
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
templates = Jinja2Templates(directory=TEMPLATES_DIR)

# Add exception handlers
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    logger.error(f"Validation error: {exc}")
    return templates.TemplateResponse("error.html", {
        "request": request,
        "error_title": "Invalid Request",
        "error_message": "The uploaded file could not be processed. Please check the file format and try again.",
        "error_details": str(exc)
    }, status_code=400)

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unexpected error: {exc}")
    logger.error(traceback.format_exc())
    return templates.TemplateResponse("error.html", {
        "request": request,
        "error_title": "Server Error",
        "error_message": "An unexpected error occurred. Please try again later.",
        "error_details": str(exc) if app.debug else "Contact support if the problem persists."
    }, status_code=500)

# Create DNS analyzer instance
dns_analyzer = DNSAnalyzer()

# Ensure uploads directory exists
UPLOADS_DIR = os.path.join(STATIC_DIR, "uploads")
os.makedirs(UPLOADS_DIR, exist_ok=True)

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Serve the main upload page"""
    return templates.TemplateResponse("index.html", {"request": request})

def validate_domain(domain: str) -> bool:
    """Validate domain name to prevent injection attacks"""
    # RFC compliant domain validation
    domain_pattern = re.compile(
        r'^(?:[a-zA-Z0-9]'
        r'(?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)*'
        r'[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?$'
    )
    
    # Additional checks
    if not domain or len(domain) > 253:
        return False
    
    # Check for dangerous characters
    dangerous_chars = ['<', '>', '"', "'", '\\', '|', '&', ';', '`', '$', '(', ')', '{', '}', '[', ']']
    if any(char in domain for char in dangerous_chars):
        return False
    
    return bool(domain_pattern.match(domain))

@app.post("/analyze-single", response_class=HTMLResponse)
async def analyze_single_domain(request: Request):
    """Analyze a single domain entered by the user"""
    try:
        # Get form data
        form_data = await request.form()
        domain = form_data.get('domain', '').strip()
        
        if not domain:
            raise HTTPException(status_code=400, detail="No domain provided")
        
        # Clean domain name
        domain = domain.replace('http://', '').replace('https://', '').replace('www.', '').strip('/')
        
        # SECURITY: Validate domain format
        if not validate_domain(domain):
            raise HTTPException(status_code=400, detail="Invalid domain format. Please enter a valid domain name.")
        
        logger.info(f"Analyzing single domain: {domain}")
        
        # Analyze the domain
        result = await dns_analyzer.analyze_domain(domain)
        results = [result]
        
        # Generate summary
        summary = dns_analyzer.generate_summary(results)
        
        # Create analysis response
        analysis_response = AnalysisResponse(
            results=results,
            summary=summary,
            timestamp=datetime.now()
        )
        
        # Generate Excel report
        report_filename = await generate_excel_report(analysis_response)
        
        # Render results page
        return templates.TemplateResponse("results.html", {
            "request": request,
            "results": results,
            "summary": summary,
            "report_filename": report_filename,
            "timestamp": analysis_response.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error analyzing single domain: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error analyzing domain: {str(e)}")

def parse_multipart_data(body: bytes, boundary: str):
    """Parse multipart form data manually to avoid library issues"""
    
    # Clean boundary - remove quotes if present
    boundary = boundary.strip('"').strip("'")
    
    # Convert boundary to bytes if it's a string
    if isinstance(boundary, str):
        boundary_bytes = boundary.encode()
    else:
        boundary_bytes = boundary
    
    # Debug logging
    logger.info(f"Original boundary: {repr(boundary)}")
    logger.info(f"Boundary as bytes: {repr(boundary_bytes)}")
    logger.info(f"Body length: {len(body)} bytes")
    
    # Log first part of body for debugging
    logger.info(f"Body start: {repr(body[:100])}")
    
    # Find all boundary occurrences
    boundary_marker = b'--' + boundary_bytes
    logger.info(f"Looking for boundary marker: {repr(boundary_marker)}")
    
    parts = body.split(boundary_marker)
    
    files = {}
    
    logger.info(f"Found {len(parts)} parts after splitting by boundary")
    
    for i, part in enumerate(parts):
        logger.info(f"Part {i}: {repr(part[:200]) if len(part) > 200 else repr(part)}")
        
        if not part or part.strip() in (b'', b'--', b'--\r\n', b'--\n'):
            logger.info(f"Skipping empty part {i}")
            continue
            
        logger.info(f"Processing part {i}, length: {len(part)} bytes")
        
        # Find the double newline that separates headers from content
        header_end = -1
        sep_used = None
        for sep in (b'\r\n\r\n', b'\n\n'):
            if sep in part:
                header_end = part.find(sep)
                headers = part[:header_end]
                content = part[header_end + len(sep):]
                sep_used = sep
                break
        
        if header_end == -1:
            logger.info(f"No header separator found in part {i}, trying to find any headers")
            # Try to find just headers without content
            if b'\r\n' in part:
                headers = part.split(b'\r\n')[1] if len(part.split(b'\r\n')) > 1 else part
                content = b''
            elif b'\n' in part:
                headers = part.split(b'\n')[1] if len(part.split(b'\n')) > 1 else part
                content = b''
            else:
                logger.info(f"No recognizable structure in part {i}")
                continue
        
        # Parse headers
        headers_str = headers.decode('utf-8', errors='ignore')
        logger.info(f"Headers for part {i}: {repr(headers_str)}")
        logger.info(f"Content length for part {i}: {len(content)}")
        
        # Clean up content - remove trailing CRLF and boundary markers
        if content:
            content = content.rstrip(b'\r\n').rstrip(b'--').rstrip(b'\r\n')
            logger.info(f"Cleaned content length for part {i}: {len(content)}")
        
        # Look for file field with more flexible matching
        if ('name="file"' in headers_str or "name='file'" in headers_str or 
            'name=file' in headers_str):
            logger.info(f"Found file field in part {i}")
            
            # Extract filename with multiple patterns
            filename = None
            for pattern in [r'filename="([^"]*)"', r"filename='([^']*)'", r'filename=([^;\s\r\n]+)']:
                filename_match = re.search(pattern, headers_str)
                if filename_match:
                    filename = filename_match.group(1).strip()
                    logger.info(f"Extracted filename using pattern {pattern}: {filename}")
                    break
            
            if filename:
                logger.info(f"Found file: {filename}, content length: {len(content)}")
                files['file'] = {
                    'filename': filename,
                    'content': content
                }
            else:
                logger.info(f"File field found but no filename extracted from: {repr(headers_str)}")
                # Store it anyway if we have content
                if content:
                    files['file'] = {
                        'filename': 'uploaded_file',
                        'content': content
                    }
                    logger.info(f"Stored file without filename, content length: {len(content)}")
                
    logger.info(f"Final extracted files: {list(files.keys())}")
    return files

@app.post("/upload", response_class=HTMLResponse)
async def upload_file(request: Request):
    """Handle file upload using raw body parsing"""
    
    try:
        # Get content type and boundary
        content_type = request.headers.get('content-type', '')
        
        if not content_type.startswith('multipart/form-data'):
            raise HTTPException(status_code=400, detail="Invalid content type. Expected multipart/form-data.")
        
        # Extract boundary from content type
        boundary_match = re.search(r'boundary=([^;]+)', content_type)
        if not boundary_match:
            raise HTTPException(status_code=400, detail="No boundary found in content type.")
        
        boundary = boundary_match.group(1).strip('"')
        
        # Read raw body
        body = await request.body()
        
        if not body:
            raise HTTPException(status_code=400, detail="No file data received.")
        
        # Parse multipart data manually
        files = parse_multipart_data(body, boundary)
        
        if 'file' not in files:
            raise HTTPException(status_code=400, detail="No file found in upload.")
        
        file_data = files['file']
        filename = file_data['filename']
        file_content = file_data['content']
        
        if not filename:
            raise HTTPException(status_code=400, detail="No filename provided.")
        
        # Validate file type
        allowed_extensions = {'.xlsx', '.xls', '.csv'}
        file_extension = os.path.splitext(filename)[1].lower()
        
        if file_extension not in allowed_extensions:
            raise HTTPException(
                status_code=400, 
                detail="Invalid file type. Please upload Excel (.xlsx, .xls) or CSV files."
            )
        
        # Validate file size
        max_size = 10 * 1024 * 1024  # 10MB
        if len(file_content) > max_size:
            raise HTTPException(status_code=400, detail="File too large. Maximum size is 10MB.")
        
        if not file_content:
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")
        
        # Parse file based on extension
        domains = []
        try:
            if file_extension == '.csv':
                # Read CSV file
                df = pd.read_csv(BytesIO(file_content))
            else:
                # Read Excel file
                df = pd.read_excel(BytesIO(file_content))
        except Exception as e:
            raise HTTPException(
                status_code=400, 
                detail=f"Error reading file: {str(e)}. Please ensure the file is not corrupted."
            )
        
        # Validate that 'Domain' column exists
        if 'Domain' not in df.columns:
            available_columns = ', '.join(df.columns.tolist())
            raise HTTPException(
                status_code=400, 
                detail=f"File must contain a 'Domain' column with domain names to analyze. Available columns: {available_columns}"
            )
        
        # Extract domains
        domains = df['Domain'].dropna().astype(str).tolist()
        domains = [domain.strip() for domain in domains if domain.strip() and domain.strip().lower() != 'nan']
        
        # SECURITY: Validate all domains
        validated_domains = []
        for domain in domains:
            # Clean domain
            domain = domain.replace('http://', '').replace('https://', '').replace('www.', '').strip('/')
            if validate_domain(domain):
                validated_domains.append(domain)
            else:
                logger.warning(f"Skipping invalid domain: {domain}")
        
        domains = validated_domains
        
        if not domains:
            raise HTTPException(
                status_code=400, 
                detail="No valid domains found in the uploaded file."
            )
        
        # SECURITY: Limit number of domains to prevent DoS
        max_domains = 1000
        if len(domains) > max_domains:
            raise HTTPException(
                status_code=400,
                detail=f"Too many domains. Maximum allowed is {max_domains} domains per request."
            )
        
        logger.info(f"Processing {len(domains)} domains from {filename}")
        
        # Analyze domains
        results = await dns_analyzer.analyze_domains(domains)
        
        # Generate summary
        summary = dns_analyzer.generate_summary(results)
        
        # Create analysis response
        analysis_response = AnalysisResponse(
            results=results,
            summary=summary,
            timestamp=datetime.now()
        )
        
        # Generate Excel report
        report_filename = await generate_excel_report(analysis_response)
        
        # Render results page with download link
        return templates.TemplateResponse("results.html", {
            "request": request,
            "results": results,
            "summary": summary,
            "report_filename": report_filename,
            "timestamp": analysis_response.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        })
        
    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        logger.error(f"Unexpected error processing uploaded file: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Unexpected error processing file: {str(e)}")

@app.get("/download/{filename}")
async def download_report(filename: str):
    """Download generated Excel report"""
    # SECURITY: Prevent path traversal attacks
    # Only allow alphanumeric, hyphens, underscores, and .xlsx extension
    if not re.match(r'^[a-zA-Z0-9_-]+\.xlsx$', filename):
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    # SECURITY: Use Path to safely resolve the file path
    upload_dir = Path(UPLOADS_DIR).resolve()
    file_path = (upload_dir / filename).resolve()
    
    # SECURITY: Ensure the resolved path is within the uploads directory
    if not str(file_path).startswith(str(upload_dir)):
        raise HTTPException(status_code=403, detail="Access denied")
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

async def generate_excel_report(analysis_response: AnalysisResponse) -> str:
    """Generate Excel report from analysis results"""
    
    # Create DataFrame from results
    data = []
    for result in analysis_response.results:
        data.append({
            'Domain': result.domain,
            'MX_Records': result.mx_records,
            'SPF_Record': result.spf_record,
            'DMARC_Record': result.dmarc_record,
            'DMARC_Policy': result.dmarc_policy,
            'Has_DMARC_Policy': result.has_dmarc_policy,
            'DMARC_RUF_Email': result.dmarc_ruf_email,
            'Has_Website': result.has_website,
            'A_Record': result.a_record,
            'WWW_CNAME': result.www_cname,
            'WWW_Points_To_Main': result.www_points_to_main,
            'Tech_Contact_Email': result.tech_contact_email,
            'Error': result.error
        })
    
    df = pd.DataFrame(data)
    
    # Generate unique filename
    timestamp = analysis_response.timestamp.strftime("%Y%m%d_%H%M%S")
    filename = f"dns_analysis_report_{timestamp}_{uuid.uuid4().hex[:8]}.xlsx"
    file_path = os.path.join(UPLOADS_DIR, filename)
    
    # Create Excel file with formatting
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DNS Analysis', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['DNS Analysis']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add conditional formatting for DMARC policy
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import FormulaRule
        
        # Find the Has_DMARC_Policy column
        has_policy_col = None
        for i, cell in enumerate(worksheet[1]):
            if cell.value == 'Has_DMARC_Policy':
                has_policy_col = i + 1
                break
        
        if has_policy_col:
            # Green fill for True values
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            # Red fill for False values  
            red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            
            # Apply conditional formatting
            for row in range(2, len(df) + 2):  # Skip header row
                cell = worksheet.cell(row=row, column=has_policy_col)
                if cell.value == True:
                    cell.fill = green_fill
                elif cell.value == False:
                    cell.fill = red_fill
    
    logger.info(f"Generated Excel report: {filename}")
    return filename

@app.post("/api/analyze")
async def analyze_domains_api(analysis_request: AnalysisRequest):
    """API endpoint to analyze domains from JSON request"""
    
    try:
        domains = [domain.strip() for domain in analysis_request.domains if domain.strip()]
        
        if not domains:
            raise HTTPException(status_code=400, detail="No valid domains provided")
        
        logger.info(f"API: Processing {len(domains)} domains")
        
        # Analyze domains
        results = await dns_analyzer.analyze_domains(domains)
        
        # Generate summary
        summary = dns_analyzer.generate_summary(results)
        
        # Create analysis response
        analysis_response = AnalysisResponse(
            results=results,
            summary=summary,
            timestamp=datetime.now()
        )
        
        return analysis_response
        
    except Exception as e:
        logger.error(f"API error analyzing domains: {e}")
        raise HTTPException(status_code=500, detail=f"Error analyzing domains: {str(e)}")

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.now()}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)